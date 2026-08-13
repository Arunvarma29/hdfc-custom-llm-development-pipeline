from io import BytesIO
from typing import Any

from openpyxl import load_workbook


def parse_xlsx(file_stream: BytesIO) -> list[dict[str, Any]]:
    """
    Parse an XLSX file into a list of dictionaries.

    The first row is treated as the header.
    Each subsequent row becomes one record.
    """

    file_stream.seek(0)

    workbook = load_workbook(
        filename=file_stream,
        read_only=True,
        data_only=True,
    )

    try:
        worksheet = workbook.active

        rows = worksheet.iter_rows(values_only=True)

        headers = next(rows, None)

        if headers is None:
            return []

        headers = [
            str(header).strip() if header is not None else ""
            for header in headers
        ]

        records: list[dict[str, Any]] = []

        for row in rows:
            record = {
                header: value
                for header, value in zip(headers, row)
                if header
            }

            records.append(record)

        return records

    finally:
        workbook.close()